# -*- coding: utf-8 -*-
# @Author: iwiniwin
# @Date:   2018-02-01 10:33:33
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

import wx
import os
import random
import json
import random
import time
import re
import codecs
from functions import *
from threading import Thread
import threading
import thread
import xlwt
import xlrd
import openpyxl
from openpyxl.styles import Font, Fill

phone_pattern = re.compile(r'(?<!\d)1[3-9]\d{9}(?!\d)')

email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,3}')

xls_suffix = ".xls"
xlsx_suffix = ".xlsx"


config = {"filter_chinese" : True, "filter_filename" : "config"}
if os.path.isfile("config.json"):
	config = json.load(open("config.json", 'r'))

def dump(*args):
	# print(args)
	pass

# http://www.ityouknow.com/python/2019/12/29/python-excel-103.html  参考文档

def extract(path):
	(filepath,tempfilename) = os.path.split(path)
	(filename,extension) = os.path.splitext(tempfilename)
	phones, emails = read_book(path, extension)
	if extension == xlsx_suffix:
		return write_book2(u'提取结果-' + filename + extension, phones, emails)
	else:
		return write_book(u'提取结果-' + filename + extension, phones, emails)

# 4053 5030

# 9659
# 15974
# 3986
# 7197
def read_book(path, extension):
	

	phones = []
	emails = []

	if extension == xlsx_suffix:
		inwb = openpyxl.load_workbook(path) 
		for sheet in inwb:
			read_sheet2(sheet, phones, emails)
	else:
		wb = xlrd.open_workbook(path)
		for i in range(0, wb.nsheets):
			read_sheet(wb.sheet_by_index(i), phones, emails)

	dump(len(phones))
	dump(len(emails))

	# phones = sorted(set(phones), key = phones.index)  # 用于过滤重复元素
	# emails = sorted(set(emails), key = emails.index)  # 用于过滤重复元素
	phones = {}.fromkeys(phones).keys()
	emails = {}.fromkeys(emails).keys()

	dump(len(phones))
	dump(len(emails))

	return phones, emails

def read_sheet(sh, phones, emails):
	# 处理一个表

	dump(u"sheet %s 共 %d 行 %d 列" % (sh.name, sh.nrows, sh.ncols))

	for r in range(0, sh.nrows):
		for c in range(0, sh.ncols):
			t = str(sh.cell_value(r, c))
			p = phone_pattern.findall(t)
			if len(p) > 0:
				phones.extend(p)
			p = email_pattern.findall(t)
			if len(p) > 0:
					emails.extend(p)

def read_sheet2(sh, phones, emails):
	# 处理一个表
	dump(u"sheet %s 共 %d 行 %d 列" % (sh.title, sh.max_row, sh.max_column))
	for row in sh.iter_rows(min_row=1, min_col=1, max_row=sh.max_row, max_col=sh.max_column):
	    for cell in row:
			t = str(cell.value) #读文件
			p = phone_pattern.findall(t)
			if len(p) > 0:
				phones.extend(p)
			p = email_pattern.findall(t)
			if len(p) > 0:
				emails.extend(p)

def write_book(path, phones, emails):
	# 创建 xls 文件对象
	wb = xlwt.Workbook()

	sh = wb.add_sheet(u'提取结果')

	# 设置列宽
	sh.col(0).width = 256 * 20
	sh.col(1).width = 256 * 40

	# 设置行高
	tall_style = xlwt.easyxf('font:height 280')  # 36pt
	sh.row(0).set_style(tall_style)
	

	# 然后按照位置来添加数据,第一个参数是行，第二个参数是列
	sh.write(0, 0, u'手机号码')
	sh.write(0, 1, u'邮箱')

	index = 1
	for phone in phones:
		sh.row(index).set_style(tall_style)
		sh.write(index, 0, phone)
		index = index + 1

	index = 1
	for email in emails:
		sh.row(index).set_style(tall_style)
		sh.write(index, 1, email)
		index = index + 1


	# 最后保存文件即可
	try:
		wb.save(path)
	except Exception, err:
		return err
	return True

def write_book2(path, phones, emails):
	book = openpyxl.Workbook()
	sheet = book.active

	sheet['A1'] = u'手机号码'
	sheet['B1'] = u'邮箱'

	sheet.column_dimensions['A'].width = 20
	sheet.column_dimensions['B'].width = 40
	row_height = 16
	sheet.row_dimensions[1].height = row_height

	index = 2
	for phone in phones:
		sheet.row_dimensions[index].height = row_height
		sheet['A' + str(index)] = phone
		index = index + 1

	index = 2
	for email in emails:
		sheet.row_dimensions[index].height = row_height
		sheet['B' + str(index)] = email
		index = index + 1

	try:
		book.save(path)
	except Exception, err:
		return err
	return True

class ExtractThread(Thread):
	def __init__(self, fileNames):
		super(ExtractThread, self).__init__()
		self.fileNames = fileNames
		self.start()

	def run(self):
		for fileName in self.fileNames:
			ret = extract(fileName)
			if ret != True:
				wx._button.SetLabel("拖拽区域")
				wx.MessageBox("提取失败 ：" + str(ret).replace('u\'','\'').decode("unicode-escape"), 'Error', wx.YES_DEFAULT | wx.ICON_ERROR)
				return
		wx._button.SetLabel("拖拽区域")
		wx.MessageBox("提取成功 :  文件生成在当前程序所在目录",'Info',wx.OK|wx.ICON_INFORMATION)

	def get_result(self):
		threading.Thread.join(self) # 等待线程执行完毕
		try:
			return self.ret
		except Exception, err:
			return err

class FileDropTarget(wx.FileDropTarget):  
	def __init__(self, window):  
		wx.FileDropTarget.__init__(self)  
		self.window = window  

	def OnDropFiles(self,  x,  y, fileNames):  
		wx._button.SetLabel("正在提取相关信息，这需要一些时间，请稍候。。。")
		t = ExtractThread(fileNames)
		# ret = t.get_result()
		# if ret == True:
		# 	wx.MessageBox("提取成功",'Info',wx.OK|wx.ICON_INFORMATION)
		# else:
		# 	wx.MessageBox("提取失败 ：" + str(ret), 'Error', wx.YES_DEFAULT | wx.ICON_ERROR)
		return True

class MyFrame(wx.Frame):

    def __init__(self):
		wx.Frame.__init__(self, None, title=u'excel提取工具2.0',size=(480,320))
		self.Center()
		# self.icon = wx.Icon('icon.ico', wx.BITMAP_TYPE_ICO)
		# self.SetIcon(self.icon)
		panel = wx.Panel(self)
		button = wx.Button(panel, label = u"拖拽区域", size = (480, 290))
		wx._button = button

		self.dropTarget = FileDropTarget(self)  
		self.SetDropTarget(self.dropTarget)

		


if __name__ == "__main__":
	app = wx.App(True)

	frm = MyFrame()
	frm.Show()
	app.MainLoop()
